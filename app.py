from flask import Flask, render_template, request, jsonify, session, redirect, url_for
from flask_cors import CORS
import requests
import os
from datetime import timedelta, datetime
import hashlib
import psycopg2
from psycopg2.extras import RealDictCursor
from dotenv import load_dotenv
import uuid
from threading import Thread
import time

# Carregar variáveis de ambiente
dotenv_path = os.path.join(os.path.dirname(__file__), '.env')
load_dotenv(dotenv_path)

app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=2)
CORS(app)

# Configurações do Banco de Dados PostgreSQL
DB_CONFIG = {
    'host': os.getenv('HOST_DW'),
    'database': os.getenv('DBNAME_DW'),
    'user': os.getenv('USER_DW'),
    'password': os.getenv('PASS_DW'),
    'port': os.getenv('PORT_DW', '5432'),
    'options': '-c search_path=apontador_horas,public'
}

# URL do webhook do n8n
N8N_WEBHOOK_URL = "https://n8n.bookerbrasil.com/webhook/9d8f9a85-c21d-4aed-bd52-124af0d116c3/chat"

# Sistema de alertas - POR USUÁRIO
alertas_por_usuario = {}  # {usuario: [alertas]}

def get_db_connection():
    """Cria uma conexão com o banco de dados PostgreSQL"""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        return conn
    except Exception as e:
        print(f"❌ Erro ao conectar no banco: {e}")
        return None

def hash_senha(senha):
    """Gera hash SHA-256 da senha"""
    return hashlib.sha256(senha.encode()).hexdigest()

def contar_tarefas_ativas_por_usuario():
    """Verifica tarefas ativas por usuário e retorna dicionário {usuario: quantidade}"""
    conn = get_db_connection()
    if not conn:
        return None
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute("""
            SELECT 
                usuario,
                COUNT(*) as total
            FROM apontador_horas.v_tarefas_ativas 
            WHERE status = 'em_andamento'
            GROUP BY usuario
        """)
        resultados = cursor.fetchall()
        
        # Converter para dicionário {usuario: total}
        tarefas_por_usuario = {}
        for row in resultados:
            tarefas_por_usuario[row['usuario']] = row['total']
        
        return tarefas_por_usuario
    except Exception as e:
        print(f"❌ Erro ao verificar tarefas ativas: {e}")
        return None
    finally:
        conn.close()

def adicionar_alerta_usuario(usuario, mensagem):
    """Adiciona um alerta para um usuário específico com timestamp de expiração"""
    global alertas_por_usuario
    
    timestamp = datetime.now().strftime('%H:%M:%S')
    alerta = {
        'id': str(uuid.uuid4()),
        'mensagem': mensagem,
        'timestamp': timestamp,
        'hora': datetime.now().strftime('%H:%M'),
        'criado_em': datetime.now()  # Timestamp completo para verificar expiração
    }
    
    # Inicializar lista de alertas do usuário se não existir
    if usuario not in alertas_por_usuario:
        alertas_por_usuario[usuario] = []
    
    alertas_por_usuario[usuario].append(alerta)
    print(f"⚠️ ALERTA para {usuario}: {mensagem} às {timestamp}")

def limpar_alertas_expirados():
    """Remove alertas com mais de 2 horas"""
    global alertas_por_usuario
    agora = datetime.now()
    tempo_expiracao = timedelta(hours=2)
    
    usuarios_para_limpar = []
    
    for usuario, alertas in alertas_por_usuario.items():
        # Filtrar alertas não expirados
        alertas_validos = [
            alerta for alerta in alertas
            if (agora - alerta['criado_em']) < tempo_expiracao
        ]
        
        if len(alertas_validos) < len(alertas):
            removidos = len(alertas) - len(alertas_validos)
            print(f"🗑️ Removidos {removidos} alerta(s) expirado(s) de {usuario}")
        
        if len(alertas_validos) > 0:
            alertas_por_usuario[usuario] = alertas_validos
        else:
            usuarios_para_limpar.append(usuario)
    
    # Remover usuários sem alertas
    for usuario in usuarios_para_limpar:
        del alertas_por_usuario[usuario]

def scheduler_verificacao_tarefas():
    """Thread que verifica tarefas ativas às 17:00 e 18:00 - POR USUÁRIO"""
    print("🕒 Scheduler de verificação de tarefas iniciado (por usuário)")
    
    ultimo_horario_executado = None  # Rastrear último horário que executou
    
    while True:
        now = datetime.now()
        hora_atual = now.strftime('%H:%M')
        
        # Verificar se é um horário agendado E se ainda não executou neste horário
        if hora_atual in ['17:00', '18:00'] and hora_atual != ultimo_horario_executado:
            print(f"🔔 Executando verificação de tarefas às {hora_atual}")
            
            tarefas_por_usuario = contar_tarefas_ativas_por_usuario()
            
            if tarefas_por_usuario is not None:
                if len(tarefas_por_usuario) == 0:
                    # Ninguém tem tarefas ativas
                    print(f"✅ Nenhum usuário com tarefas em andamento às {hora_atual}")
                else:
                    # Criar alerta individual para cada usuário com tarefas ativas
                    for usuario, total in tarefas_por_usuario.items():
                        mensagem = f"⚠️ Atenção! Você tem {total} tarefa(s) em andamento às {hora_atual}"
                        adicionar_alerta_usuario(usuario, mensagem)
                    
                    print(f"✅ {len(tarefas_por_usuario)} usuário(s) notificado(s) às {hora_atual}")
            else:
                print(f"❌ Erro ao verificar tarefas às {hora_atual}")
            
            # Marcar este horário como executado
            ultimo_horario_executado = hora_atual
            print(f"✅ Verificação concluída. Próxima verificação será em outro horário.")
        
        # Verificar a cada 30 segundos
        time.sleep(30)

def verificar_usuario(usuario, senha):
    """Verifica se usuário e senha estão corretos no banco de dados"""
    conn = get_db_connection()
    if not conn:
        return False
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute("""
            SELECT id, usuario, senha_hash, nome_completo, email, departamento, nivel, ativo
            FROM funcionarios
            WHERE usuario = %s
        """, (usuario,))
        
        user = cursor.fetchone()
        
        if not user:
            print(f"❌ Usuário '{usuario}' não encontrado")
            return None
        
        if not user['ativo']:
            print(f"❌ Usuário '{usuario}' está inativo")
            return None
        
        senha_hash = hash_senha(senha)
        
        if user['senha_hash'] == senha_hash:
            print(f"✅ Login válido para '{usuario}'")
            return user
        else:
            print(f"❌ Senha incorreta para '{usuario}'")
            return None
        
    except Exception as e:
        print(f"❌ Erro ao verificar usuário: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        conn.close()

def get_usuarios_permitidos(usuario_logado, nivel_logado):
    """
    Retorna lista de usuários que o usuário logado pode visualizar
    
    Regras:
    - funcionario, prestador de servico: apenas ele mesmo
    - coordenador, supervisor: ele mesmo + subordinados (onde ele é gestor)
    - socio, admin: todos os usuários
    """
    conn = get_db_connection()
    if not conn:
        return [usuario_logado]  # Fallback: apenas ele mesmo
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Admin e Sócio: todos os usuários
        if nivel_logado in ['admin', 'socio']:
            cursor.execute("""
                SELECT usuario FROM funcionarios WHERE ativo = TRUE
            """)
            usuarios = [r['usuario'] for r in cursor.fetchall()]
            print(f"🔓 {usuario_logado} ({nivel_logado}): acesso TOTAL a {len(usuarios)} usuários")
            return usuarios
        
        # Coordenador e Supervisor: ele mesmo + subordinados
        elif nivel_logado in ['coordenador', 'supervisor']:
            # Buscar: ele mesmo + quem tem ele como gestor no campo nome_gestor
            cursor.execute("""
                SELECT usuario 
                FROM funcionarios 
                WHERE ativo = TRUE 
                  AND (usuario = %s OR nome_gestor = %s)
            """, (usuario_logado, usuario_logado))
            
            usuarios = [r['usuario'] for r in cursor.fetchall()]
            print(f"👥 {usuario_logado} ({nivel_logado}): acesso a {len(usuarios)} usuários (ele + subordinados)")
            return usuarios
        
        # Funcionário e Prestador: apenas ele mesmo
        else:
            print(f"🔒 {usuario_logado} ({nivel_logado}): acesso RESTRITO (apenas ele mesmo)")
            return [usuario_logado]
        
    except Exception as e:
        print(f"❌ Erro ao buscar usuários permitidos: {e}")
        import traceback
        traceback.print_exc()
        return [usuario_logado]  # Fallback: apenas ele mesmo
    finally:
        conn.close()

# ========================================
# ROTAS DE AUTENTICAÇÃO
# ========================================

@app.route('/')
def index():
    if 'usuario' in session:
        return render_template('chat.html', usuario=session['usuario'])
    return redirect(url_for('login_page'))

@app.route('/login')
def login_page():
    return render_template('login.html')

@app.route('/api/login', methods=['POST'])
def login():
    dados = request.get_json()
    usuario = dados.get('usuario')
    senha = dados.get('senha')
    
    if not usuario or not senha:
        return jsonify({'success': False, 'message': 'Usuário e senha são obrigatórios'}), 400
    
    user = verificar_usuario(usuario, senha)
    
    if user:
        session_id = str(uuid.uuid4())
        
        session['usuario'] = user['usuario']
        session['usuario_id'] = user['id']
        session['nome_completo'] = user['nome_completo']
        session['nivel'] = user['nivel']
        session['departamento'] = user['departamento']
        session['session_id'] = session_id
        session.permanent = True
        
        print(f"✅ Login: {user['usuario']} | Nível: {user['nivel']} | Session ID: {session_id}")
        
        return jsonify({'success': True, 'message': 'Login realizado com sucesso'})
    
    return jsonify({'success': False, 'message': 'Usuário ou senha inválidos'}), 401

@app.route('/api/logout', methods=['POST'])
def logout():
    usuario = session.get('usuario', 'desconhecido')
    session_id = session.get('session_id', 'sem-session')
    
    print(f"🚪 Logout: {usuario} | Session ID: {session_id}")
    
    session.clear()
    return jsonify({'success': True})

@app.route('/api/usuario-info', methods=['GET'])
def usuario_info():
    """Retorna informações do usuário logado"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    return jsonify({
        'success': True,
        'usuario': session.get('usuario'),
        'nome_completo': session.get('nome_completo'),
        'nivel': session.get('nivel'),
        'departamento': session.get('departamento')
    })

@app.route('/api/alertas', methods=['GET'])
def obter_alertas():
    """Retorna alertas do usuário logado (remove expirados automaticamente)"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    usuario = session.get('usuario')
    global alertas_por_usuario
    
    # Limpar alertas expirados de todos os usuários
    limpar_alertas_expirados()
    
    # Retornar apenas alertas deste usuário
    alertas_usuario = alertas_por_usuario.get(usuario, [])
    
    # Remover campo criado_em antes de enviar (não é necessário no frontend)
    alertas_para_enviar = []
    for alerta in alertas_usuario:
        alerta_limpo = {
            'id': alerta['id'],
            'mensagem': alerta['mensagem'],
            'timestamp': alerta['timestamp'],
            'hora': alerta['hora']
        }
        alertas_para_enviar.append(alerta_limpo)
    
    return jsonify({
        'success': True,
        'alertas': alertas_para_enviar
    })

@app.route('/api/alertas/limpar', methods=['POST'])
def limpar_alertas():
    """Limpa alertas do usuário logado"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    usuario = session.get('usuario')
    global alertas_por_usuario
    
    # Limpar apenas alertas deste usuário
    if usuario in alertas_por_usuario:
        alertas_por_usuario[usuario] = []
    
    return jsonify({'success': True})

@app.route('/api/alertas/visualizar/<alerta_id>', methods=['POST'])
def marcar_visualizado(alerta_id):
    """Marca um alerta específico como visualizado (remove da lista)"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    usuario = session.get('usuario')
    global alertas_por_usuario
    
    if usuario in alertas_por_usuario:
        # Remover alerta com este ID
        alertas_por_usuario[usuario] = [
            alerta for alerta in alertas_por_usuario[usuario]
            if alerta['id'] != alerta_id
        ]
        print(f"✅ Alerta {alerta_id} marcado como visualizado por {usuario}")
    
    return jsonify({'success': True})

# ========================================
# ROTAS DE BUSCA
# ========================================

@app.route('/api/buscar-clientes', methods=['POST'])
def buscar_clientes():
    """Busca clientes por nome ou CNPJ (formatado ou somente números)"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    dados = request.get_json()
    query = dados.get('query', '').strip()
    
    if len(query) < 2:
        return jsonify({'success': True, 'clientes': []})
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Erro de conexão'}), 500
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Remove formatação do CNPJ/CPF para busca (pontos, barras, hífens)
        query_numeros = ''.join(filter(str.isdigit, query))
        
        cursor.execute("""
            SELECT num_cnpj_cpf, nom_cliente, cod_grupo_cliente, des_grupo
            FROM clientes
            WHERE LOWER(nom_cliente) LIKE LOWER(%s)
               OR LOWER(des_grupo) LIKE LOWER(%s)
               OR REPLACE(REPLACE(REPLACE(REPLACE(num_cnpj_cpf, '.', ''), '/', ''), '-', ''), ' ', '') LIKE %s
            ORDER BY 
                CASE 
                    WHEN LOWER(nom_cliente) = LOWER(%s) THEN 1
                    WHEN LOWER(nom_cliente) LIKE LOWER(%s) THEN 2
                    WHEN REPLACE(REPLACE(REPLACE(REPLACE(num_cnpj_cpf, '.', ''), '/', ''), '-', ''), ' ', '') = %s THEN 1
                    ELSE 3
                END,
                nom_cliente
            LIMIT 10
        """, (f'%{query}%', f'%{query}%', f'%{query_numeros}%', query, f'{query}%', query_numeros))
        
        clientes = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'clientes': [dict(c) for c in clientes]
        })
        
    except Exception as e:
        print(f"❌ Erro ao buscar clientes: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/buscar-tarefas', methods=['POST'])
def buscar_tarefas():
    """Busca tarefas para o usuário logado - opcionalmente filtrado por cliente"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    dados = request.get_json()
    cnpj = dados.get('cnpj', '').strip()
    usuario = session.get('usuario')
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Erro de conexão'}), 500
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Se CNPJ foi fornecido, busca apenas tarefas daquele cliente
        if cnpj:
            cursor.execute("""
                SELECT 
                    t.id,
                    t.nome_tarefa,
                    t.cod_grupo_tarefa,
                    t.prioridade,
                    t.estimativa_horas,
                    t.cnpj_cpf,
                    c.nom_cliente,
                    c.des_grupo
                FROM tarefas_colaborador t
                LEFT JOIN clientes c ON t.cnpj_cpf = c.num_cnpj_cpf
                WHERE t.cnpj_cpf = %s
                  AND (t.colaborador_1 = %s OR t.colaborador_2 = %s)
                ORDER BY 
                    CASE 
                        WHEN LOWER(t.prioridade) LIKE '%%alta%%' OR LOWER(t.prioridade) LIKE '%%p2%%' THEN 1
                        WHEN LOWER(t.prioridade) LIKE '%%média%%' OR LOWER(t.prioridade) LIKE '%%media%%' OR LOWER(t.prioridade) LIKE '%%p1%%' THEN 2
                        WHEN LOWER(t.prioridade) LIKE '%%baixa%%' OR LOWER(t.prioridade) LIKE '%%p3%%' THEN 3
                        ELSE 4
                    END,
                    t.nome_tarefa
            """, (cnpj, usuario, usuario))
        else:
            # Se CNPJ não foi fornecido, busca todas as tarefas do usuário
            cursor.execute("""
                SELECT 
                    t.id,
                    t.nome_tarefa,
                    t.cod_grupo_tarefa,
                    t.prioridade,
                    t.estimativa_horas,
                    t.cnpj_cpf,
                    c.nom_cliente,
                    c.des_grupo
                FROM tarefas_colaborador t
                LEFT JOIN clientes c ON t.cnpj_cpf = c.num_cnpj_cpf
                WHERE (t.colaborador_1 = %s OR t.colaborador_2 = %s)
                ORDER BY 
                    CASE 
                        WHEN LOWER(t.prioridade) LIKE '%%alta%%' OR LOWER(t.prioridade) LIKE '%%p2%%' THEN 1
                        WHEN LOWER(t.prioridade) LIKE '%%média%%' OR LOWER(t.prioridade) LIKE '%%media%%' OR LOWER(t.prioridade) LIKE '%%p1%%' THEN 2
                        WHEN LOWER(t.prioridade) LIKE '%%baixa%%' OR LOWER(t.prioridade) LIKE '%%p3%%' THEN 3
                        ELSE 4
                    END,
                    c.nom_cliente,
                    t.nome_tarefa
            """, (usuario, usuario))
        
        tarefas = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'tarefas': [dict(t) for t in tarefas]
        })
        
    except Exception as e:
        print(f"❌ Erro ao buscar tarefas: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

# ========================================
# ROTAS DE CONTROLE DE TAREFAS (MÚLTIPLAS)
# ========================================

@app.route('/api/iniciar-tarefa', methods=['POST'])
def iniciar_tarefa():
    """Inicia uma nova tarefa"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    dados = request.get_json()
    cnpj_cliente = dados.get('cnpj_cliente')
    nome_cliente = dados.get('nome_cliente')
    tarefa_id = dados.get('tarefa_id')
    nome_tarefa = dados.get('nome_tarefa')
    observacao = dados.get('observacao', '')  # ⭐ NOVO: Campo observação (opcional)
    usuario = session.get('usuario')
    
    if not all([cnpj_cliente, tarefa_id]):
        return jsonify({'success': False, 'message': 'Dados incompletos'}), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Erro de conexão'}), 500
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute("""
            WITH ids_resolvidos AS (
                SELECT 
                    f.id AS funcionario_id,
                    c.id AS cliente_id,
                    f.nome_completo,
                    c.nom_cliente
                FROM funcionarios f
                CROSS JOIN clientes c
                WHERE f.usuario = %s
                  AND c.num_cnpj_cpf = %s
                LIMIT 1
            )
            INSERT INTO apontamentos_horas (
                funcionario_id,
                cliente_id,
                tarefa_id,
                data_inicio,
                status,
                observacao  -- ⭐ NOVO: Campo observação
            )
            SELECT 
                funcionario_id,
                cliente_id,
                %s,
                NOW(),
                'em_andamento',
                %s  -- ⭐ NOVO: Parâmetro observação
            FROM ids_resolvidos
            RETURNING 
                id,
                TO_CHAR(data_inicio AT TIME ZONE 'America/Sao_Paulo', 'YYYY-MM-DD HH24:MI:SS') AS data_inicio_br
        """, (usuario, cnpj_cliente, tarefa_id, observacao))  # ⭐ NOVO: Adicionar observacao nos parâmetros
        
        resultado = cursor.fetchone()
        conn.commit()
        
        print(f"✅ Tarefa iniciada: {usuario} | ID: {tarefa_id} | {nome_tarefa} | Cliente: {nome_cliente} | Obs: {observacao[:50] if observacao else 'N/A'}")
        
        return jsonify({
            'success': True,
            'apontamento_id': resultado['id'],
            'data_inicio': resultado['data_inicio_br'],
            'message': 'Tarefa iniciada com sucesso'
        })
        
    except Exception as e:
        conn.rollback()
        print(f"❌ Erro ao iniciar tarefa: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/pausar-tarefa', methods=['POST'])
def pausar_tarefa():
    """Pausa uma tarefa específica"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    dados = request.get_json()
    apontamento_id = dados.get('apontamento_id')
    
    if not apontamento_id:
        return jsonify({'success': False, 'message': 'ID da tarefa não fornecido'}), 400
    
    usuario = session.get('usuario')
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Erro de conexão'}), 500
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Verificar se a tarefa pertence ao usuário
        cursor.execute("""
            SELECT a.id 
            FROM apontamentos_horas a
            INNER JOIN funcionarios f ON a.funcionario_id = f.id
            WHERE a.id = %s AND f.usuario = %s AND a.status = 'em_andamento'
        """, (apontamento_id, usuario))
        
        if not cursor.fetchone():
            return jsonify({'success': False, 'message': 'Tarefa não encontrada ou não pode ser pausada'}), 400
        
        # Pausar tarefa
        cursor.execute("""
            UPDATE apontamentos_horas
            SET status = 'pausado', atualizado_em = NOW()
            WHERE id = %s
        """, (apontamento_id,))
        
        # Criar registro de pausa
        cursor.execute("""
            INSERT INTO pausas (apontamento_id, data_pausa)
            VALUES (%s, NOW())
        """, (apontamento_id,))
        
        conn.commit()
        
        print(f"⏸️ Tarefa {apontamento_id} pausada: {usuario}")
        return jsonify({'success': True})
        
    except Exception as e:
        conn.rollback()
        print(f"❌ Erro ao pausar tarefa: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/retomar-tarefa', methods=['POST'])
def retomar_tarefa():
    """Retoma uma tarefa específica"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    dados = request.get_json()
    apontamento_id = dados.get('apontamento_id')
    
    if not apontamento_id:
        return jsonify({'success': False, 'message': 'ID da tarefa não fornecido'}), 400
    
    usuario = session.get('usuario')
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Erro de conexão'}), 500
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Verificar se a tarefa pertence ao usuário
        cursor.execute("""
            SELECT a.id 
            FROM apontamentos_horas a
            INNER JOIN funcionarios f ON a.funcionario_id = f.id
            WHERE a.id = %s AND f.usuario = %s AND a.status = 'pausado'
        """, (apontamento_id, usuario))
        
        if not cursor.fetchone():
            return jsonify({'success': False, 'message': 'Tarefa não encontrada ou não está pausada'}), 400
        
        # Fechar pausa atual
        cursor.execute("""
            UPDATE pausas
            SET data_retomada = NOW()
            WHERE apontamento_id = %s AND data_retomada IS NULL
        """, (apontamento_id,))
        
        # Retomar tarefa
        cursor.execute("""
            UPDATE apontamentos_horas
            SET status = 'em_andamento', atualizado_em = NOW()
            WHERE id = %s
        """, (apontamento_id,))
        
        conn.commit()
        
        print(f"▶️ Tarefa {apontamento_id} retomada: {usuario}")
        return jsonify({'success': True})
        
    except Exception as e:
        conn.rollback()
        print(f"❌ Erro ao retomar tarefa: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/finalizar-tarefa', methods=['POST'])
def finalizar_tarefa():
    """Finaliza uma tarefa específica"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    dados = request.get_json()
    apontamento_id = dados.get('apontamento_id')
    
    if not apontamento_id:
        return jsonify({'success': False, 'message': 'ID da tarefa não fornecido'}), 400
    
    usuario = session.get('usuario')
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Erro de conexão'}), 500
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Verificar se a tarefa pertence ao usuário
        cursor.execute("""
            SELECT a.id 
            FROM apontamentos_horas a
            INNER JOIN funcionarios f ON a.funcionario_id = f.id
            WHERE a.id = %s AND f.usuario = %s AND a.status IN ('em_andamento', 'pausado')
        """, (apontamento_id, usuario))
        
        if not cursor.fetchone():
            return jsonify({'success': False, 'message': 'Tarefa não encontrada ou já finalizada'}), 400
        
        # Fechar pausa se existir
        cursor.execute("""
            UPDATE pausas
            SET data_retomada = NOW()
            WHERE apontamento_id = %s AND data_retomada IS NULL
        """, (apontamento_id,))
        
        # Finalizar tarefa
        cursor.execute("""
            UPDATE apontamentos_horas
            SET data_fim = NOW(), 
                status = 'finalizado', 
                atualizado_em = NOW()
            WHERE id = %s
            RETURNING data_inicio, data_fim
        """, (apontamento_id,))
        
        tarefa = cursor.fetchone()
        
        # Calcular horas
        cursor.execute("""
            SELECT 
                EXTRACT(EPOCH FROM (%s - %s))/3600 AS horas_totais,
                COALESCE(
                    (SELECT SUM(EXTRACT(EPOCH FROM (
                        COALESCE(p.data_retomada, NOW()) - p.data_pausa
                    )))/3600
                    FROM pausas p
                    WHERE p.apontamento_id = %s),
                    0
                ) AS horas_pausadas
        """, (tarefa['data_fim'], tarefa['data_inicio'], apontamento_id))
        
        tempos = cursor.fetchone()
        horas_trabalhadas = tempos['horas_totais'] - tempos['horas_pausadas']
        
        conn.commit()
        
        print(f"✅ Tarefa {apontamento_id} finalizada: {usuario} | {horas_trabalhadas:.2f}h")
        return jsonify({
            'success': True,
            'horas_trabalhadas': round(horas_trabalhadas, 2)
        })
        
    except Exception as e:
        conn.rollback()
        print(f"❌ Erro ao finalizar tarefa: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/registrar-atrasado', methods=['POST'])
def registrar_atrasado():
    """Registra apontamento de horas atrasado (com data/hora passadas)"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    dados = request.get_json()
    cnpj_cliente = dados.get('cnpj_cliente')
    nome_cliente = dados.get('nome_cliente')
    tarefa_id = dados.get('tarefa_id')
    nome_tarefa = dados.get('nome_tarefa')
    data_inicio = dados.get('data_inicio')
    data_fim = dados.get('data_fim')
    observacao = dados.get('observacao', '')  # ⭐ NOVO: Campo observação (opcional)
    usuario = session.get('usuario')
    
    if not all([cnpj_cliente, tarefa_id, data_inicio, data_fim]):
        return jsonify({'success': False, 'message': 'Dados incompletos'}), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Erro de conexão'}), 500
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Validar e converter datas
        from datetime import datetime
        try:
            dt_inicio = datetime.fromisoformat(data_inicio.replace('Z', '+00:00'))
            dt_fim = datetime.fromisoformat(data_fim.replace('Z', '+00:00'))
            
            if dt_fim <= dt_inicio:
                return jsonify({'success': False, 'message': 'Data fim deve ser posterior à data início'}), 400
            
            # Converter para string no formato adequado para PostgreSQL
            data_inicio_str = dt_inicio.strftime('%Y-%m-%d %H:%M:%S')
            data_fim_str = dt_fim.strftime('%Y-%m-%d %H:%M:%S')
            
        except ValueError as e:
            return jsonify({'success': False, 'message': f'Formato de data inválido: {e}'}), 400
        
        cursor.execute("""
            WITH ids_resolvidos AS (
                SELECT 
                    f.id AS funcionario_id,
                    c.id AS cliente_id
                FROM funcionarios f
                CROSS JOIN clientes c
                WHERE f.usuario = %s
                  AND c.num_cnpj_cpf = %s
                LIMIT 1
            )
            INSERT INTO apontamentos_horas (
                funcionario_id,
                cliente_id,
                tarefa_id,
                data_inicio,
                data_fim,
                status,
                criado_em,
                atualizado_em,
                observacao  -- ⭐ NOVO: Campo observação
            )
            SELECT 
                funcionario_id,
                cliente_id,
                %s,
                %s AT TIME ZONE 'America/Sao_Paulo',
                %s AT TIME ZONE 'America/Sao_Paulo',
                'finalizado',
                NOW(),
                NOW(),
                %s  -- ⭐ NOVO: Parâmetro observação
            FROM ids_resolvidos
            RETURNING 
                id,
                EXTRACT(EPOCH FROM (data_fim - data_inicio))/3600 AS horas_trabalhadas
        """, (usuario, cnpj_cliente, tarefa_id, data_inicio_str, data_fim_str, observacao))  # ⭐ NOVO: Adicionar observacao
        
        resultado = cursor.fetchone()
        conn.commit()
        
        horas = round(resultado['horas_trabalhadas'], 2)
        
        print(f"✅ Apontamento atrasado registrado: {usuario} | Tarefa ID: {tarefa_id} | {nome_tarefa} | {horas}h | {data_inicio_str} até {data_fim_str} | Obs: {observacao[:50] if observacao else 'N/A'}")
        
        return jsonify({
            'success': True,
            'apontamento_id': resultado['id'],
            'horas_trabalhadas': horas
        })
        
    except Exception as e:
        conn.rollback()
        print(f"❌ Erro ao registrar atrasado: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/verificar-tarefas-ativas', methods=['GET'])
def verificar_tarefas_ativas():
    """Verifica TODAS as tarefas ativas do usuário"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    usuario = session.get('usuario')
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Erro de conexão'}), 500
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute("""
            SELECT 
                a.id AS apontamento_id,
                a.status,
                c.nom_cliente AS cliente_nome,
                c.num_cnpj_cpf AS cnpj,
                t.nome_tarefa AS tarefa_nome,
                a.tarefa_id,
                TO_CHAR(a.data_inicio AT TIME ZONE 'America/Sao_Paulo', 'YYYY-MM-DD HH24:MI:SS') AS data_inicio,
                CASE 
                    WHEN a.status = 'pausado' THEN 
                        (SELECT TO_CHAR(p.data_pausa AT TIME ZONE 'America/Sao_Paulo', 'YYYY-MM-DD HH24:MI:SS')
                         FROM apontador_horas.pausas p 
                         WHERE p.apontamento_id = a.id 
                           AND p.data_retomada IS NULL 
                         LIMIT 1)
                    ELSE NULL
                END AS data_pausa,
                COALESCE(
                    (SELECT SUM(EXTRACT(EPOCH FROM (
                        COALESCE(p.data_retomada, NOW()) - p.data_pausa
                    )) * 1000)
                    FROM apontador_horas.pausas p
                    WHERE p.apontamento_id = a.id),
                    0
                ) AS tempo_pausado_ms
            FROM apontador_horas.apontamentos_horas a
            INNER JOIN apontador_horas.funcionarios f ON a.funcionario_id = f.id
            INNER JOIN apontador_horas.clientes c ON a.cliente_id = c.id
            INNER JOIN apontador_horas.tarefas_colaborador t ON a.tarefa_id = t.id
            WHERE f.usuario = %s
              AND a.status IN ('em_andamento', 'pausado')
            ORDER BY a.data_inicio DESC
        """, (usuario,))
        
        tarefas = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'tarefas': [dict(t) for t in tarefas]
        })
        
    except Exception as e:
        print(f"❌ Erro ao verificar tarefas ativas: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

# ========================================
# ROTAS DE RELATÓRIOS (COM CONTROLE DE ACESSO)
# ========================================

@app.route('/api/relatorio-tempo', methods=['POST'])
def relatorio_tempo():
    """Retorna relatório de tempo decorrido por atividades"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    usuario_logado = session.get('usuario')
    nivel_logado = session.get('nivel')
    
    dados = request.get_json()
    filtros = {
        'ano': dados.get('ano'),
        'mes': dados.get('mes'),
        'departamento': dados.get('departamento'),
        'funcionario': dados.get('funcionario'),
        'grupo': dados.get('grupo'),
        'tarefa': dados.get('tarefa')
    }
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Erro de conexão'}), 500
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # 🔐 CONTROLE DE ACESSO: Obter usuários permitidos
        usuarios_permitidos = get_usuarios_permitidos(usuario_logado, nivel_logado)
        
        # Construir query dinâmica com filtros
        where_clauses = ["a.status = 'finalizado'"]
        params = []
        
        # 🔐 FILTRO OBRIGATÓRIO: Usuários permitidos
        where_clauses.append(f"f.usuario = ANY(%s)")
        params.append(usuarios_permitidos)
        
        if filtros['ano']:
            where_clauses.append("EXTRACT(YEAR FROM a.data_inicio AT TIME ZONE 'America/Sao_Paulo') = %s")
            params.append(filtros['ano'])
        
        if filtros['mes']:
            where_clauses.append("EXTRACT(MONTH FROM a.data_inicio AT TIME ZONE 'America/Sao_Paulo') = %s")
            params.append(filtros['mes'])
        
        if filtros['departamento'] and filtros['departamento'] != 'Todos':
            where_clauses.append("f.departamento = %s")
            params.append(filtros['departamento'])
        
        if filtros['funcionario'] and filtros['funcionario'] != 'Todos':
            # Verificar se o usuário filtrado está na lista permitida
            if filtros['funcionario'] in usuarios_permitidos:
                where_clauses.append("f.usuario = %s")
                params.append(filtros['funcionario'])
            else:
                # Se tentar filtrar por um usuário não permitido, retorna vazio
                return jsonify({'success': True, 'dados': {}})
        
        if filtros['grupo'] and filtros['grupo'] != 'Todos':
            where_clauses.append("c.cod_grupo_cliente = %s::INTEGER")
            params.append(filtros['grupo'])
        
        if filtros['tarefa'] and filtros['tarefa'] != 'Todos':
            where_clauses.append("t.cod_grupo_tarefa = %s")
            params.append(filtros['tarefa'])
        
        where_sql = " AND ".join(where_clauses)
        
        # Query principal
        query = f"""
            SELECT 
                c.des_grupo AS grupo_empresa,
                c.nom_cliente AS nome_cliente,
                f.nome_completo AS funcionario,
                t.cod_grupo_tarefa,
                gt.nome_grupo_tarefa AS nome_tarefa,
                COALESCE(
                    ROUND(
                        EXTRACT(EPOCH FROM SUM(a.horas_trabalhadas)) / 3600, 
                        2
                    ), 
                    0
                ) AS horas_totais
            FROM apontador_horas.apontamentos_horas a
            INNER JOIN apontador_horas.funcionarios f ON a.funcionario_id = f.id
            INNER JOIN apontador_horas.clientes c ON a.cliente_id = c.id
            INNER JOIN apontador_horas.tarefas_colaborador t ON a.tarefa_id = t.id
            INNER JOIN apontador_horas.grupo_tarefas gt ON t.cod_grupo_tarefa = gt.cod_grupo_tarefa
            WHERE {where_sql}
            GROUP BY 
                c.des_grupo,
                c.nom_cliente,
                f.nome_completo,
                t.cod_grupo_tarefa,
                gt.nome_grupo_tarefa
            ORDER BY 
                c.des_grupo,
                c.nom_cliente,
                f.nome_completo,
                gt.nome_grupo_tarefa
        """
        
        cursor.execute(query, params)
        resultados = cursor.fetchall()
        
        # Organizar dados hierarquicamente
        dados_hierarquicos = {}
        
        for row in resultados:
            grupo = row['grupo_empresa'] or 'SEM GRUPO'
            cliente = row['nome_cliente']
            funcionario = row['funcionario']
            tarefa = row['nome_tarefa']
            horas = float(row['horas_totais'])
            
            if grupo not in dados_hierarquicos:
                dados_hierarquicos[grupo] = {}
            
            if cliente not in dados_hierarquicos[grupo]:
                dados_hierarquicos[grupo][cliente] = {}
            
            if funcionario not in dados_hierarquicos[grupo][cliente]:
                dados_hierarquicos[grupo][cliente][funcionario] = {}
            
            dados_hierarquicos[grupo][cliente][funcionario][tarefa] = horas
        
        print(f"📊 Relatório gerado: {usuario_logado} ({nivel_logado}) - {len(resultados)} registros")
        
        return jsonify({
            'success': True,
            'dados': dados_hierarquicos
        })
        
    except Exception as e:
        print(f"❌ Erro ao gerar relatório: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/dashboard-dados', methods=['POST'])
def dashboard_dados():
    """Retorna dados para o dashboard"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    dados = request.get_json()
    filtros = {
        'ano': dados.get('ano'),
        'mes': dados.get('mes')
    }
    
    usuario = session.get('usuario')
    nivel = session.get('nivel', 'funcionario')
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Erro de conexão'}), 500
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Obter usuários permitidos (controle de acesso)
        usuarios_permitidos = get_usuarios_permitidos(usuario, nivel)
        
        # Construir filtros de data
        where_clauses = ["a.status = 'finalizado'"]
        params = []
        
        if filtros['ano']:
            where_clauses.append("EXTRACT(YEAR FROM a.data_inicio AT TIME ZONE 'America/Sao_Paulo') = %s")
            params.append(filtros['ano'])
        
        if filtros['mes']:
            where_clauses.append("EXTRACT(MONTH FROM a.data_inicio AT TIME ZONE 'America/Sao_Paulo') = %s")
            params.append(filtros['mes'])
        
        # Filtro de usuários permitidos
        if usuarios_permitidos:
            placeholders = ','.join(['%s'] * len(usuarios_permitidos))
            where_clauses.append(f"f.usuario IN ({placeholders})")
            params.extend(usuarios_permitidos)
        
        where_sql = " AND ".join(where_clauses)
        
        # ==== RESUMO GERAL ====
        cursor.execute(f"""
            SELECT 
                COUNT(*) as total_tarefas
            FROM apontador_horas.apontamentos_horas a
            INNER JOIN apontador_horas.funcionarios f ON a.funcionario_id = f.id
            WHERE {where_sql}
        """, params)
        resumo_result = cursor.fetchone()
        total_tarefas = resumo_result['total_tarefas'] if resumo_result else 0
        
        # Para simplificar, consideramos todas como concluídas (status = 'finalizado')
        tarefas_concluidas = total_tarefas
        tarefas_nao_concluidas = 0  # Poderia buscar de outra tabela se necessário
        
        resumo = {
            'total': total_tarefas,
            'concluidas': tarefas_concluidas,
            'nao_concluidas': tarefas_nao_concluidas
        }
        
        # ==== TAREFAS POR DEPARTAMENTO ====
        cursor.execute(f"""
            SELECT 
                f.departamento,
                COUNT(*) as quantidade
            FROM apontador_horas.apontamentos_horas a
            INNER JOIN apontador_horas.funcionarios f ON a.funcionario_id = f.id
            WHERE {where_sql}
            GROUP BY f.departamento
            ORDER BY quantidade DESC
        """, params)
        por_departamento = [dict(row) for row in cursor.fetchall()]
        
        # ==== TAREFAS POR GRUPO DE ATIVIDADE ====
        cursor.execute(f"""
            SELECT 
                gt.nome_grupo_tarefa as grupo_atividade,
                COUNT(*) as quantidade
            FROM apontador_horas.apontamentos_horas a
            INNER JOIN apontador_horas.funcionarios f ON a.funcionario_id = f.id
            INNER JOIN apontador_horas.tarefas_colaborador t ON a.tarefa_id = t.id
            INNER JOIN apontador_horas.grupo_tarefas gt ON t.cod_grupo_tarefa = gt.cod_grupo_tarefa
            WHERE {where_sql}
            GROUP BY gt.nome_grupo_tarefa
            ORDER BY quantidade DESC
        """, params)
        por_grupo_atividade = [dict(row) for row in cursor.fetchall()]
        
        # ==== LINHA DO TEMPO MENSAL ====
        # Se mês específico foi selecionado, mostrar dia a dia
        if filtros['mes']:
            cursor.execute(f"""
                SELECT 
                    TO_CHAR(a.data_inicio AT TIME ZONE 'America/Sao_Paulo', 'DD/MM') as dia,
                    COUNT(*) as quantidade
                FROM apontador_horas.apontamentos_horas a
                INNER JOIN apontador_horas.funcionarios f ON a.funcionario_id = f.id
                WHERE {where_sql}
                GROUP BY TO_CHAR(a.data_inicio AT TIME ZONE 'America/Sao_Paulo', 'DD/MM'),
                         DATE(a.data_inicio AT TIME ZONE 'America/Sao_Paulo')
                ORDER BY DATE(a.data_inicio AT TIME ZONE 'America/Sao_Paulo')
            """, params)
        else:
            # Se não selecionou mês, mostrar mês a mês
            cursor.execute(f"""
                SELECT 
                    TO_CHAR(a.data_inicio AT TIME ZONE 'America/Sao_Paulo', 'Mon') as dia,
                    COUNT(*) as quantidade
                FROM apontador_horas.apontamentos_horas a
                INNER JOIN apontador_horas.funcionarios f ON a.funcionario_id = f.id
                WHERE {where_sql}
                GROUP BY TO_CHAR(a.data_inicio AT TIME ZONE 'America/Sao_Paulo', 'Mon'),
                         EXTRACT(MONTH FROM a.data_inicio AT TIME ZONE 'America/Sao_Paulo')
                ORDER BY EXTRACT(MONTH FROM a.data_inicio AT TIME ZONE 'America/Sao_Paulo')
            """, params)
        
        tempo_mensal = [dict(row) for row in cursor.fetchall()]
        
        return jsonify({
            'success': True,
            'resumo': resumo,
            'por_departamento': por_departamento,
            'por_grupo_atividade': por_grupo_atividade,
            'tempo_mensal': tempo_mensal
        })
        
    except Exception as e:
        print(f"❌ Erro ao gerar dashboard: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/exportar-relatorio-excel', methods=['POST'])
def exportar_relatorio_excel():
    """Exporta relatório para Excel"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    usuario_logado = session.get('usuario')
    nivel_logado = session.get('nivel')
    
    dados = request.get_json()
    filtros = {
        'ano': dados.get('ano'),
        'mes': dados.get('mes'),
        'departamento': dados.get('departamento'),
        'funcionario': dados.get('funcionario'),
        'grupo': dados.get('grupo'),
        'tarefa': dados.get('tarefa')
    }
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Erro de conexão'}), 500
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from datetime import datetime
        
        # Buscar usuários permitidos
        usuarios_permitidos = get_usuarios_permitidos(usuario_logado, nivel_logado)
        
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Construir query (mesma lógica do relatório web)
        where_clauses = ["a.status = 'finalizado'"]
        params = []
        
        # Filtro obrigatório: apenas usuários permitidos
        where_clauses.append("f.usuario = ANY(%s)")
        params.append(usuarios_permitidos)
        
        if filtros['ano']:
            where_clauses.append("EXTRACT(YEAR FROM a.data_inicio AT TIME ZONE 'America/Sao_Paulo') = %s")
            params.append(filtros['ano'])
        
        if filtros['mes']:
            where_clauses.append("EXTRACT(MONTH FROM a.data_inicio AT TIME ZONE 'America/Sao_Paulo') = %s")
            params.append(filtros['mes'])
        
        if filtros['departamento'] and filtros['departamento'] != 'Todos':
            where_clauses.append("f.departamento = %s")
            params.append(filtros['departamento'])
        
        if filtros['funcionario'] and filtros['funcionario'] != 'Todos':
            if filtros['funcionario'] in usuarios_permitidos:
                where_clauses.append("f.usuario = %s")
                params.append(filtros['funcionario'])
            else:
                return jsonify({'success': False, 'message': 'Acesso negado'}), 403
        
        if filtros['grupo'] and filtros['grupo'] != 'Todos':
            where_clauses.append("c.cod_grupo_cliente = %s::INTEGER")
            params.append(filtros['grupo'])
        
        if filtros['tarefa'] and filtros['tarefa'] != 'Todos':
            where_clauses.append("t.cod_grupo_tarefa = %s")
            params.append(filtros['tarefa'])
        
        where_sql = " AND ".join(where_clauses)
        
        # Query principal
        query = f"""
            SELECT 
                c.des_grupo AS grupo_empresa,
                c.nom_cliente AS nome_cliente,
                f.nome_completo AS funcionario,
                gt.nome_grupo_tarefa AS nome_tarefa,
                COALESCE(
                    ROUND(
                        EXTRACT(EPOCH FROM SUM(a.horas_trabalhadas)) / 3600, 
                        2
                    ), 
                    0
                ) AS horas_totais
            FROM apontador_horas.apontamentos_horas a
            INNER JOIN apontador_horas.funcionarios f ON a.funcionario_id = f.id
            INNER JOIN apontador_horas.clientes c ON a.cliente_id = c.id
            INNER JOIN apontador_horas.tarefas_colaborador t ON a.tarefa_id = t.id
            INNER JOIN apontador_horas.grupo_tarefas gt ON t.cod_grupo_tarefa = gt.cod_grupo_tarefa
            WHERE {where_sql}
            GROUP BY 
                c.des_grupo,
                c.nom_cliente,
                f.nome_completo,
                gt.nome_grupo_tarefa
            ORDER BY 
                c.des_grupo,
                c.nom_cliente,
                f.nome_completo,
                gt.nome_grupo_tarefa
        """
        
        cursor.execute(query, params)
        resultados = cursor.fetchall()
        
        if len(resultados) == 0:
            return jsonify({'success': False, 'message': 'Nenhum dado para exportar'}), 400
        
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório de Horas"
        
        # Estilos
        header_fill = PatternFill(start_color="FFD500", end_color="FFD500", fill_type="solid")
        header_font = Font(bold=True, color="3F3F41", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = "RELATÓRIO DE APONTAMENTO DE HORAS - BOOKER BRASIL"
        title_cell.font = Font(bold=True, size=14, color="3F3F41")
        title_cell.fill = PatternFill(start_color="FFD500", end_color="FFD500", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Informações do filtro
        linha = 3
        ws[f'A{linha}'] = f"Gerado por: {session.get('nome_completo', usuario_logado)}"
        ws[f'A{linha}'].font = Font(bold=True)
        linha += 1
        
        ws[f'A{linha}'] = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws[f'A{linha}'].font = Font(bold=True)
        linha += 1
        
        # Filtros aplicados
        if filtros['ano']:
            ws[f'A{linha}'] = f"Ano: {filtros['ano']}"
            linha += 1
        if filtros['mes']:
            meses = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
            ws[f'A{linha}'] = f"Mês: {meses[int(filtros['mes'])]}"
            linha += 1
        
        linha += 1  # Linha em branco
        
        # Cabeçalhos
        headers = ['Grupo de Empresas', 'Cliente', 'Funcionário', 'Tarefa', 'Horas']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=linha, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        linha += 1
        
        # Função para converter horas decimais em formato HH:MM
        def converter_horas_para_tempo(horas_decimal):
            """Converte horas decimais (ex: 2.5) para formato de tempo HH:MM (ex: 02:30)"""
            horas = int(horas_decimal)
            minutos = int((horas_decimal - horas) * 60)
            return f"{horas:02d}:{minutos:02d}"
        
        # Dados
        total_geral = 0
        for row in resultados:
            ws.cell(row=linha, column=1, value=row['grupo_empresa'] or 'SEM GRUPO')
            ws.cell(row=linha, column=2, value=row['nome_cliente'])
            ws.cell(row=linha, column=3, value=row['funcionario'])
            ws.cell(row=linha, column=4, value=row['nome_tarefa'])
            
            # Converter horas decimais para formato HH:MM
            horas_decimal = float(row['horas_totais'])
            horas_formatadas = converter_horas_para_tempo(horas_decimal)
            
            horas_cell = ws.cell(row=linha, column=5, value=horas_formatadas)
            horas_cell.alignment = Alignment(horizontal='center')
            
            total_geral += horas_decimal
            
            # Aplicar bordas
            for col in range(1, 6):
                ws.cell(row=linha, column=col).border = border
            
            linha += 1
        
        # Total geral
        linha += 1
        total_cell = ws.cell(row=linha, column=4)
        total_cell.value = "TOTAL GERAL:"
        total_cell.font = Font(bold=True, size=12)
        total_cell.alignment = Alignment(horizontal='right')
        
        # Converter total geral para formato HH:MM
        total_formatado = converter_horas_para_tempo(total_geral)
        
        total_value_cell = ws.cell(row=linha, column=5)
        total_value_cell.value = total_formatado
        total_value_cell.font = Font(bold=True, size=12)
        total_value_cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
        total_value_cell.alignment = Alignment(horizontal='center')
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 12
        
        # Salvar em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Gerar nome do arquivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        nome_arquivo = f"relatorio_horas_{timestamp}.xlsx"
        
        print(f"📊 Excel exportado: {usuario_logado} - {len(resultados)} registros")
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nome_arquivo
        )
        
    except Exception as e:
        print(f"❌ Erro ao exportar Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/filtros-relatorio', methods=['GET'])
def filtros_relatorio():
    """Retorna opções disponíveis para filtros (com controle de acesso)"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    usuario_logado = session.get('usuario')
    nivel_logado = session.get('nivel')
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Erro de conexão'}), 500
    
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # 🔐 CONTROLE DE ACESSO: Obter usuários permitidos
        usuarios_permitidos = get_usuarios_permitidos(usuario_logado, nivel_logado)
        
        # Departamentos (dos usuários permitidos)
        cursor.execute("""
            SELECT DISTINCT departamento 
            FROM apontador_horas.funcionarios 
            WHERE ativo = TRUE AND usuario = ANY(%s)
            ORDER BY departamento
        """, (usuarios_permitidos,))
        departamentos = [r['departamento'] for r in cursor.fetchall()]
        
        # Funcionários (somente os permitidos)
        cursor.execute("""
            SELECT usuario, nome_completo 
            FROM apontador_horas.funcionarios 
            WHERE ativo = TRUE AND usuario = ANY(%s)
            ORDER BY nome_completo
        """, (usuarios_permitidos,))
        funcionarios = cursor.fetchall()
        
        # Grupos de clientes (mantém todos)
        cursor.execute("""
            SELECT DISTINCT cod_grupo_cliente, des_grupo 
            FROM apontador_horas.clientes 
            WHERE cod_grupo_cliente IS NOT NULL 
            ORDER BY des_grupo
        """)
        grupos_clientes = cursor.fetchall()
        
        # Grupos de tarefas (mantém todos)
        cursor.execute("""
            SELECT cod_grupo_tarefa, nome_grupo_tarefa 
            FROM apontador_horas.grupo_tarefas 
            ORDER BY cod_grupo_tarefa
        """)
        grupos_tarefas = cursor.fetchall()
        
        print(f"🔍 Filtros carregados: {usuario_logado} ({nivel_logado}) - {len(funcionarios)} funcionários disponíveis")
        
        return jsonify({
            'success': True,
            'departamentos': departamentos,
            'funcionarios': [dict(f) for f in funcionarios],
            'grupos_clientes': [dict(g) for g in grupos_clientes],
            'grupos_tarefas': [dict(g) for g in grupos_tarefas]
        })
        
    except Exception as e:
        print(f"❌ Erro ao buscar filtros: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

# ========================================
# ROTA DE CHAT (mantém compatibilidade)
# ========================================

@app.route('/api/chat', methods=['POST'])
def chat():
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Não autenticado'}), 401
    
    dados = request.get_json()
    mensagem = dados.get('mensagem')
    usuario = session.get('usuario')
    nome_completo = session.get('nome_completo')
    usuario_id = session.get('usuario_id')
    session_id = session.get('session_id')
    
    if not session_id:
        session_id = str(uuid.uuid4())
        session['session_id'] = session_id
    
    if not mensagem:
        return jsonify({'success': False, 'message': 'Mensagem vazia'}), 400
    
    try:
        payload = {
            'chatInput': mensagem,
            'usuario': usuario,
            'nome_completo': nome_completo,
            'usuario_id': usuario_id,
            'sessionId': session_id
        }
        
        response = requests.post(N8N_WEBHOOK_URL, json=payload, timeout=30)
        
        if response.status_code != 200:
            return jsonify({
                'success': False,
                'message': f'Erro no servidor do chatbot (HTTP {response.status_code})'
            }), 500
        
        data = response.json()
        
        resposta_bot = None
        if isinstance(data, dict):
            resposta_bot = (
                data.get('output') or
                data.get('text') or
                data.get('response') or
                data.get('resposta') or
                data.get('message')
            )
            
            if not resposta_bot and 'data' in data:
                if isinstance(data['data'], dict):
                    resposta_bot = data['data'].get('output') or data['data'].get('text')
                elif isinstance(data['data'], str):
                    resposta_bot = data['data']
        elif isinstance(data, str):
            resposta_bot = data
        
        if not resposta_bot:
            resposta_bot = 'Desculpe, recebi uma resposta em formato inesperado do servidor.'
        
        return jsonify({
            'success': True,
            'resposta': resposta_bot
        })
        
    except requests.exceptions.Timeout:
        return jsonify({
            'success': False,
            'message': 'O chatbot demorou muito para responder. Tente novamente.'
        }), 500
    except Exception as e:
        print(f"❌ Erro no chat: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Erro ao processar resposta: {str(e)}'
        }), 500

if __name__ == '__main__':
    # Teste de conexão ao iniciar
    print("🔍 Testando conexão com banco de dados...")
    conn = get_db_connection()
    if conn:
        print("✅ Conexão com PostgreSQL estabelecida!")
        conn.close()
    else:
        print("⚠️ AVISO: Não foi possível conectar ao banco de dados!")
    
    # Iniciar scheduler de verificação de tarefas em thread separada
    scheduler_thread = Thread(target=scheduler_verificacao_tarefas, daemon=True)
    scheduler_thread.start()
    print("✅ Scheduler de verificação de tarefas iniciado (17:00 e 18:00)")
    
    app.run(debug=True, host='0.0.0.0', port=5000)